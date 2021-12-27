# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

from openpyxl import Workbook
import openpyxl
from tkinter import *
import tkinter as tk
def main():

    T = []
    new = []
    wb_obj = openpyxl.load_workbook("Crew.xlsx")
    sheet_obj = wb_obj.active
    for row in sheet_obj.values:
        for value in row:
            new.append(value)
        T.append(new)
        new = []

    rows = []
    for i in range(sheet_obj.max_row):
        cols = []
        for j in range(sheet_obj.max_column):
            e = Entry(relief=GROOVE)

            e.grid(row=i, column=j, sticky=NSEW)

            e.insert(END, T[i][j])

            cols.append(e)

        rows.append(cols)

    mainloop()

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
